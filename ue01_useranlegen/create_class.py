import logging
import argparse
import random
import sys
from logging.handlers import RotatingFileHandler

from openpyxl import load_workbook

logger = logging.getLogger("create_user_logger")

'''
    - ein Bash-Script2 mit allen notwendigen Schritten/Befehlen zum Erzeugen der Benutzer3
    – ein Bash-Script zum Löschen der angelegten Benutzer
    – eine Liste mit Usernamen und Passwörtern zum Verteilen an die unterrichtenden Lehrer
    – ein Logfile mit sinnvollen Angaben
'''
'''
    – Einlesen (in Liste/Generator oder ähnliches)
    – Aufbereiten/Nachbessern:
        ∗ PW erzeugen11
        ∗ bei echten Usernamen: Korrektur des Namens (Umlaute), doppelte Namen, etc. ∗ Ordnernamen
    – Ausgabe
'''


def read_excel(file_path: str):
    """
    Reads an Excel-File and yields the output per line
    :param file_path: Excel file
    :return: tuple of class, classroom, teacher
    """

    logger.debug(f"reading from: {file_path}")

    try:
        workbook = load_workbook(file_path, read_only=True)
    except Exception:
        logger.error("Error: File " + file_path + " not Found!")
        raise Exception("File \"" + file_path + "\" not Found!")

    worksheet = workbook[workbook.sheetnames[0]]

    for row in worksheet.iter_rows(min_row=1):
        class_name = row[0].value
        class_room = row[1].value
        class_teacher = row[2].value

        if class_name is not None and class_room is not None and class_teacher is not None:
            logger.debug(f"yielding: {class_name} {class_room} {class_teacher}")
            yield class_name, class_room, class_teacher


def write_class_script(data, filename: str):
    """
    Writes data consisting of user entries in a given file
    :param data: generator object (home_directory, username, main_group, groups)
    :param filename: file to be written
    :return: bash script for user creation
    """

    groups = "cdrom,plugdev,sambashare"

    with open(filename, "w") as script, open("passwords.txt", "w") as passwords:
        for entry in data:
            home_directory, username, main_group, password = entry
            script.write("useradd "
                         " -d " + home_directory +
                         " -c " + username +
                         " -m " +
                         " -g " + main_group +
                         " -G " + groups +
                         " -s /bin/bash " + username +
                         "\n"
                         )
            logger.debug(f"writing to script: \"Klasse:  {username}, Passwort: {password}")

            script.write("echo " + username + ":" + password + " | chpasswd\n\n")
            passwords.write("Klasse:   " + username + "\nPasswort: " + password.replace("\\", "") + "\n\n")

        for additional in create_additional_users():
            home_directory, username, main_group = additional
            script.write("useradd "
                         " -d " + home_directory +
                         " -c " + username +
                         " -m " +
                         " -g " + main_group +
                         " -G " + groups +
                         " -s /bin/bash " + username +
                         "\n"
                         )
            logger.debug("writing additional users")


def create_class_users(filename: str):
    """
    generates user data from a given excel file
    :param filename: excel file
    :return: yield (home_directory, class_name, class_name, groups)
    """
    logger.debug(f"generating class users from: {filename}")

    home_directory = "/home/klassen/"
    class_names = []

    excel = read_excel(filename)
    excel.__next__()
    for row in excel:
        class_name = "k" + row[0].lower()
        if class_name in class_names:
            logger.error("user '" + class_name + "' already exists!")
            raise Exception("user '" + class_name + "' already exists!")
        else:
            class_names.append(class_name)
            logger.debug(f"yielding class user: {class_name}")
            yield (home_directory + class_name, class_name, class_name
                   , generate_password(class_name, str(row[1]), str(row[2])))


def create_additional_users():
    yield "/home/lehrer", "lehrer", "lehrer"
    yield "/home/lehrer", "seminar", "seminar"


def generate_password(name, room, teacher):
    """
    Das Passwort besteht aus KZRZJZ
        ∗ K(lasse)
        ∗ Z(ufall): ein zufälliges Zeichen aus !%&(),._-=^#5
        ∗ R(aum), ohne dem B für Beamer
        ∗ J(ahrgangsvorstand)
    :return:
    """
    logger.debug(f"generating password for: {name} {room} {teacher}")

    rand_chars = ["!", "%", "&", "(", ")", ",", ".", "_", "-", "=", "^", "#"]

    return (name +
            "\\" + random.choice(rand_chars)
            + room
            + "\\" + random.choice(rand_chars)
            + teacher
            + "\\" + random.choice(rand_chars)
            )


# def create_password_file():
#     pass


def parse_args():
    """
    The necessary commands to make the command line tools usable
    :return:
    """
    create_logger()

    parser = argparse.ArgumentParser()
    parser.add_argument("filename")
    parser.add_argument('-v', "--verbosity", help="increase output verbosity", action="store_true")
    parser.add_argument("-q", "--quiet", help="decrease output verbosity", action="store_true")

    args = parser.parse_args()

    if args.verbosity:
        print("verbosity turned on")
        logger.setLevel(logging.DEBUG)
    elif args.quiet:
        logger.setLevel(logging.NOTSET)

    if args.filename:
        write_class_script(create_class_users(args.filename), "user_script.sh")


def create_logger():
    rot_file_handler = RotatingFileHandler('create_class.log', maxBytes=10_000, backupCount=5)
    stream_handler = logging.StreamHandler(sys.stdout)
    logger.addHandler(rot_file_handler)
    logger.addHandler(stream_handler)


if __name__ == '__main__':
    parse_args()

# TODO: – bei bereits existierenden Benutzern sollte eine Fehlermeldung erfolgen????????????
# TODO: – die Bash-Scripts sollten bei Problemen abbrechen
