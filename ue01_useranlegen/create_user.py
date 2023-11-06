"""
Als Login-Name wird der Familienname verwendet:
– Nur Kleinbuchstaben.
– Alle Sonderzeichen (z.B. Umlaute) werden ersetzt (zB aus Ä wird ae, aus ß wird ss).
– Akzente müssen entfernt werden.
– Leerzeichen im Namen werden durch Unterstrich ersetzt.
– Bei mehreren gleichen Namen werden die Zahlen 1,2,3, ... angehängt (Beispiel: maier, maier1, maier2, ...).
• Das Passwort sollte zufällig gewählt werden.
• Für folgende Teilaufgaben sollten Tests enthalten sein:
– Sonderzeichen im Benutzernamen – Doppelte Benutzernamen
"""
import argparse
import logging
import random
import sys
import unicodedata
import re
from logging.handlers import RotatingFileHandler

import openpyxl
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

from create_class import generate_password

logger = logging.getLogger("create_user_logger")


def read_excel(file_path: str):
    """
    Reads an Excel-File and yields the output per line
    :param file_path: Excel file
    :return: tuple of class, classroom, teacher
    """

    logger.debug(f"reading from: {file_path}")

    try:
        workbook = load_workbook(file_path, read_only=True)
    except Exception:
        logger.error("Error: File " + file_path + " not Found!")
        raise Exception("File \"" + file_path + "\" not Found!")

    worksheet = workbook[workbook.sheetnames[0]]

    for row in worksheet.iter_rows(min_row=1):
        first_name = row[0].value
        last_name = row[1].value
        group_name = row[2].value
        class_name = row[3].value

        if first_name is not None and last_name is not None and group_name is not None:
            logger.debug(f"yielding: {first_name} {last_name} {group_name}")
            yield first_name, last_name, group_name, class_name
        else:
            logger.error(f"malformed user entry for: {row}")


def create_users(filename: str):
    """
    reads user data from excel file
    yields user data and password
    :param filename: excel file
    :return:
    """
    logger.debug(f"generating users from: {filename}")

    last_names = []
    special_char_map = {ord('ä'): 'ae', ord('ü'): 'ue', ord('ö'): 'oe', ord('ß'): 'ss',
                        ord('Ä'): 'Ae', ord('Ü'): 'Ue', ord('Ö'): 'Oe'}

    excel = read_excel(filename)
    excel.__next__()
    for row in excel:
        first_name = shave_marks(row[0].replace(" ", "_").translate(special_char_map))
        last_name = shave_marks(row[1].replace(" ", "_").translate(special_char_map))
        group_name = row[2].lower()
        class_name = row[3].lower() if row[3] else None

        name_pattern = "^[A-Za-z][A-Za-z0-9._-]+$"
        if not re.search(name_pattern, first_name) or not re.search(name_pattern, last_name):
            continue

        last_name = convert_multiple_last_names(last_name, last_names)

        user_name = last_name.lower()

        logger.debug(f"yielding class user: {first_name}")
        yield from create_user(class_name, first_name, group_name, last_name, user_name)


def create_user(class_name, first_name, group_name, last_name, user_name):
    """
    yields given user data and defines home directory
    :param class_name:
    :param first_name:
    :param group_name:
    :param last_name:
    :param user_name:
    :return: (home_directory, first_name, last_name, class_name, main_group_name, additional_groups, user_name, password)

    >>> user = [entry for entry in create_user("5cn", "tobias", "5cn", "hernandez_perez", "tobi")]
    >>> user = user.pop()[:-1] # remove password because random
    >>> user
    ('/home/klassen/5cn/hernandez_perez/', 'tobias', 'hernandez_perez', '5cn', '5cn', '5cn', 'tobi')
    """
    if class_name:
        home_directory = f"/home/klassen/{class_name}/{last_name}/"
    else:
        home_directory = f"/home/lehrer/{last_name}/"
    yield (home_directory,  # home directory
           first_name,      # first name
           last_name,       # last name
           class_name,      # class name
           group_name,      # main group name
           class_name,      # additional groups
           user_name,       # user name
           generate_password(first_name, last_name, str(class_name)))   # password


def convert_multiple_last_names(last_name, last_names):
    """
    in case of a reoccurring last name
    appends iterating number
    :param last_name:
    :param last_names:
    :return:

    >>> last_names = ["zainzinger","breunig", "wagner"]
    >>> convert_multiple_last_names("breunig", last_names)
    'breunig1'
    >>> convert_multiple_last_names("breunig", last_names)
    'breunig2'
    >>> convert_multiple_last_names("meier", last_names)
    'meier'
    """
    if last_name in last_names:
        logger.debug("lastname '" + last_name + "' already exists -> appending number")
        last_names.append(last_name)
        last_name = last_name + str(last_names.count(last_name) - 1)
    else:
        last_names.append(last_name)
    return last_name


def shave_marks(txt):
    """
    Remove all diacritic marks

    >>> shave_marks("î lövé sëw")
    'i love sew'
    >>> shave_marks("abcdefghijklmnopqrstuvwxyz")
    'abcdefghijklmnopqrstuvwxyz'
    >>> shave_marks(" ")
    ' '
    """
    norm_txt = unicodedata.normalize('NFD', txt)
    shaved = ''.join(c for c in norm_txt
                     if not unicodedata.combining(c))
    return unicodedata.normalize('NFC', shaved)


def write_excel(filename, data):
    """
    Write User Data to an Excel File
    :param filename: Excel file to write to
    :param data: User Data
    :return:
    """
    wb = openpyxl.Workbook()
    counter = 1

    ws = wb.active
    create_titles(ws)
    for user in data:
        counter += 1
        home_directory, first_name, last_name, class_name, main_group, groups, username, password = user
        ws[f'A{counter}'].value = first_name
        ws[f'B{counter}'].value = last_name
        ws[f'C{counter}'].value = class_name
        ws[f'D{counter}'].value = username
        ws[f'E{counter}'].value = password.replace('\\', '')

        style_row(counter, ws)

    autofit_columns(ws)

    wb.save(filename)


def write_txt(filename, data):
    """
    writes the users and their passwords into a txt file
    :param filename:
    :param data:
    :return:
    """
    with open(filename, "w") as passwords:
        for entry in data:
            home_directory, first_name, last_name, class_name, main_group, groups, username, password = entry
            passwords.write(f"Name:     {first_name} {last_name}\n" +
                            f"Username: {username}\n" + username + "\n" +
                            "Passwort:  " + password.replace("\\", "") + "\n\n"
                            )


def style_row(counter, ws):
    """
    applies styles to a single row of cells
        - an alternating background color
        - a thin border on all sides
    :param counter: counter used for alternation
    :param ws: Worksheet
    :return:
    """
    fill = PatternFill(start_color="BBBBBB", end_color="BBBBBB", fill_type="solid")
    thin_border = Border(
        top=Side(style='thin'),
        left=Side(style='thin'),
        right=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for cell in ws[counter]:
        if counter % 2 == 0:
            cell.fill = fill
        cell.border = thin_border


def autofit_columns(ws):
    """
    automatically adjusts the column widths to fit the contents
    :param ws: Worksheet
    :return:
    """
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width


def create_titles(ws):
    """
    creates titles of columns and styles them
        - First Name
        - Last Name
        - Class
        - Login Name
        - Password
    all titles are in bold and have medium borders left, right and bottom
    :param ws: Worksheet
    :return:
    """
    ws.title = "Users"

    a1 = ws['A1']
    a1.value = "First Name"
    a1.font = Font(bold=True)

    b1 = ws['B1']
    b1.value = "Last Name"
    b1.font = Font(bold=True)

    c1 = ws['C1']
    c1.value = "Class"
    c1.font = Font(bold=True)

    d1 = ws['D1']
    d1.value = "Login Name"
    d1.font = Font(bold=True)

    e1 = ws['E1']
    e1.value = "Password"
    e1.font = Font(bold=True)

    medium_border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        bottom=Side(style='medium')
    )
    for cell in ws[1]:
        cell.border = medium_border


def write_user_script(filename: str, data):
    """
    Writes data consisting of user entries in a given file
    :param data: generator object (home_directory, username, main_group, groups)
    :param filename: file to be written
    :return: bash script for user creation
    """

    with open(filename, "w") as script:
        for user in data:
            home_directory, first_name, last_name, class_name, main_group, groups, username, password = user
            script.write("useradd "
                         f" -d {home_directory}" +
                         f" -c \"{first_name} {last_name}\"" +
                         " -m " +
                         f" -g {main_group}" +
                         f" -G cdrom,plugdev,sambashare,{groups}" +
                         " -s /bin/bash " +
                         username +
                         "\n"
                         )
            logger.debug(f"writing to script: \"Klasse:  {username}, Passwort: {password}")

            script.write(f"echo {username}:\"{password}\" | chpasswd\n\n")


def parse_args():
    """
    The necessary commands to make the command line tools usable
    :return:
    """
    create_logger()

    parser = argparse.ArgumentParser()
    parser.add_argument("filename")
    parser.add_argument("-e", "--excel", help="set the output method to excel (default=txt)", action="store_true")
    parser.add_argument('-v', "--verbosity", help="increase output verbosity", action="store_true")
    parser.add_argument("-q", "--quiet", help="decrease output verbosity", action="store_true")

    args = parser.parse_args()

    if args.verbosity:
        print("verbosity turned on")
        logger.setLevel(logging.DEBUG)
    elif args.quiet:
        logger.setLevel(logging.NOTSET)

    if args.filename:
        write_user_script("new_user_script.sh", create_users(args.filename))
        if args.excel:
            write_excel("users.xlsx", create_users(args.filename))
        else:
            write_txt("user_passwords.txt", create_users(args.filename))


def create_logger():
    """
    creates the variables necessary for logging
    :return:
    """
    rot_file_handler = RotatingFileHandler('create_user.log', maxBytes=10_000, backupCount=5)
    stream_handler = logging.StreamHandler(sys.stdout)
    logger.addHandler(rot_file_handler)
    logger.addHandler(stream_handler)


if __name__ == '__main__':
    parse_args()

    # write_excel("users.xlsx", create_users("Namen.xlsx"))
    # write_user_script("new_user_script.sh", create_users("Namen.xlsx"))
    print("done")
